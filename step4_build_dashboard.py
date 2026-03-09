# ============================================================
#  STEP 4 — Build the Excel Dashboard
#  Reads the clean data and creates a formatted Excel file
#  with 4 sheets:
#    Sheet 1 – Student Records   (all cleaned data)
#    Sheet 2 – Summary Stats     (KPI cards + pie chart)
#    Sheet 3 – Subject Analysis  (breakdown + bar chart)
#    Sheet 4 – Attendance Report (ranked by attendance)
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Load clean data ───────────────────────────────────────
print("📂 Loading clean data...")
df = pd.read_csv("clean_student_data.csv")

# ── Create a new Excel workbook ───────────────────────────
wb = Workbook()


# ============================================================
#  HELPER FUNCTIONS  (reusable formatting shortcuts)
# ============================================================

def make_header_cell(ws, cell_address, text,
                     bg_color="2C3E50", text_color="FFFFFF",
                     font_size=11, bold=True):
    """Style a cell as a header (dark background, white text)."""
    cell = ws[cell_address]
    cell.value     = text
    cell.font      = Font(name="Arial", size=font_size,
                          bold=bold, color=text_color)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    side = Side(style="thin", color="FFFFFF")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def make_data_cell(ws, cell_address, value,
                   bg_color="FFFFFF", align="center", bold=False):
    """Style a regular data cell."""
    cell = ws[cell_address]
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, bold=bold)
    cell.fill      = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    side = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def set_column_width(ws, col_letter, width):
    """Set the width of a column."""
    ws.column_dimensions[col_letter].width = width


def set_row_height(ws, row_number, height):
    """Set the height of a row."""
    ws.row_dimensions[row_number].height = height


# ============================================================
#  SHEET 1 — Student Records
# ============================================================
print("🔨 Building Sheet 1: Student Records...")

ws1 = wb.active
ws1.title = "Student Records"
ws1.sheet_view.showGridLines = False  # hide grey gridlines
ws1.freeze_panes = "A3"               # freeze top 2 rows while scrolling

# ── Title row ─────────────────────────────────────────────
ws1.merge_cells("A1:I1")
make_header_cell(ws1, "A1", "📚  Student Data Dashboard",
                 bg_color="1A252F", text_color="F1C40F", font_size=16)
set_row_height(ws1, 1, 40)

# ── Column headers (row 2) ────────────────────────────────
column_headers = [
    ("A", "ID",           10),
    ("B", "Name",         22),
    ("C", "Grade Level",  13),
    ("D", "Subject",      16),
    ("E", "Score",         9),
    ("F", "Letter Grade", 14),
    ("G", "Performance",  18),
    ("H", "Attendance %", 14),
    ("I", "Test Avg",     11),
]

for col_letter, header_text, col_width in column_headers:
    make_header_cell(ws1, f"{col_letter}2", header_text)
    set_column_width(ws1, col_letter, col_width)
set_row_height(ws1, 2, 26)

# ── Colour maps for grades and performance ────────────────
performance_colors = {
    "Excellent":         "C6EFCE",   # green
    "Good":              "FFEB9C",   # yellow
    "Average":           "FFD8A8",   # orange
    "Needs Improvement": "FFBDBD",   # red
}

grade_colors = {
    "A": "C6EFCE",
    "B": "DAEEF3",
    "C": "FFEB9C",
    "D": "FFD8A8",
    "F": "FFBDBD",
}

# ── Data rows ─────────────────────────────────────────────
for row_num, (_, row) in enumerate(df.iterrows(), start=3):
    # Alternate row background for readability
    bg = "F7F9FC" if row_num % 2 == 0 else "FFFFFF"

    # Calculate average of the 3 tests
    test_avg = round((row["test1"] + row["test2"] + row["test3"]) / 3, 1)

    row_data = [
        ("A", row["student_id"],  bg,                                          "center"),
        ("B", row["name"],        bg,                                          "left"),
        ("C", row["grade_level"], bg,                                          "center"),
        ("D", row["subject"],     bg,                                          "center"),
        ("E", int(row["score"]),  bg,                                          "center"),
        ("F", row["letter_grade"],grade_colors.get(row["letter_grade"], bg),   "center"),
        ("G", row["performance"], performance_colors.get(row["performance"],bg),"center"),
        ("H", int(row["attendance_pct"]), bg,                                  "center"),
        ("I", test_avg,           bg,                                          "center"),
    ]

    for col_letter, value, cell_bg, align in row_data:
        make_data_cell(ws1, f"{col_letter}{row_num}",
                       value, bg_color=cell_bg, align=align)
    set_row_height(ws1, row_num, 18)

# Colour scale on Score column — red → yellow → green
last_data_row = 2 + len(df)
ws1.conditional_formatting.add(
    f"E3:E{last_data_row}",
    ColorScaleRule(
        start_type="min",       start_color="FF9999",
        mid_type="percentile",  mid_value=50,  mid_color="FFEB9C",
        end_type="max",         end_color="63BE7B"
    )
)


# ============================================================
#  SHEET 2 — Summary Statistics
# ============================================================
print("🔨 Building Sheet 2: Summary Statistics...")

ws2 = wb.create_sheet("Summary Statistics")
ws2.sheet_view.showGridLines = False

# ── Title ─────────────────────────────────────────────────
ws2.merge_cells("A1:F1")
make_header_cell(ws2, "A1", "📊  Summary Statistics",
                 bg_color="1A252F", text_color="F1C40F", font_size=16)
set_row_height(ws2, 1, 40)

# ── KPI Cards ─────────────────────────────────────────────
scores     = df["score"]
pass_count = (scores >= 60).sum()
pass_rate  = f"{pass_count / len(df) * 100:.1f}%"
avg_att    = f"{df['attendance_pct'].mean():.1f}%"

kpi_data = [
    ("A", "Total Records",  len(df),               "2980B9"),
    ("B", "Average Score",  f"{scores.mean():.1f}", "27AE60"),
    ("C", "Highest Score",  int(scores.max()),      "8E44AD"),
    ("D", "Lowest Score",   int(scores.min()),      "E74C3C"),
    ("E", "Pass Rate",      pass_rate,              "F39C12"),
    ("F", "Avg Attendance", avg_att,                "16A085"),
]

for col_letter, label, value, color in kpi_data:
    # Label cell
    lc = ws2[f"{col_letter}3"]
    lc.value     = label
    lc.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    lc.fill      = PatternFill("solid", start_color=color)
    lc.alignment = Alignment(horizontal="center", vertical="center",
                              wrap_text=True)

    # Value cell (big number)
    ws2.merge_cells(f"{col_letter}4:{col_letter}5")
    vc = ws2[f"{col_letter}4"]
    vc.value     = value
    vc.font      = Font(name="Arial", size=20, bold=True, color=color)
    vc.alignment = Alignment(horizontal="center", vertical="center")

    set_column_width(ws2, col_letter, 17)

for r in [3, 4, 5]:
    set_row_height(ws2, r, 24)

# ── Grade Distribution Table ──────────────────────────────
ws2["A7"].value = "Grade Distribution"
ws2["A7"].font  = Font(name="Arial", size=13, bold=True, color="2C3E50")
set_row_height(ws2, 7, 22)

for ci, h in enumerate(["Grade", "Count", "Percentage"], 1):
    make_header_cell(ws2, ws2.cell(row=8, column=ci).coordinate, h)
set_row_height(ws2, 8, 22)

grade_counts = df["letter_grade"].value_counts().sort_index()
for ri, (grade, count) in enumerate(grade_counts.items(), start=9):
    bg  = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
    pct = f"{count / len(df) * 100:.1f}%"
    make_data_cell(ws2, f"A{ri}", grade, bg)
    make_data_cell(ws2, f"B{ri}", count, bg)
    make_data_cell(ws2, f"C{ri}", pct,   bg)
    set_row_height(ws2, ri, 20)

# ── Pie Chart: Grade Distribution ────────────────────────
# Write chart source data in columns H and I
ws2["H8"].value = "Grade"
ws2["I8"].value = "Count"
for ri, (grade, count) in enumerate(grade_counts.items(), start=9):
    ws2.cell(row=ri, column=8).value = f"Grade {grade}"
    ws2.cell(row=ri, column=9).value = int(count)

pie = PieChart()
pie.title  = "Grade Distribution"
pie.style  = 10
pie.width  = 15
pie.height = 12

data_ref = Reference(ws2, min_col=9, min_row=8,
                     max_row=8 + len(grade_counts))
cats_ref = Reference(ws2, min_col=8, min_row=9,
                     max_row=8 + len(grade_counts))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
ws2.add_chart(pie, "A12")


# ============================================================
#  SHEET 3 — Subject Analysis
# ============================================================
print("🔨 Building Sheet 3: Subject Analysis...")

ws3 = wb.create_sheet("Subject Analysis")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:F1")
make_header_cell(ws3, "A1", "📈  Subject-wise Analysis",
                 bg_color="1A252F", text_color="F1C40F", font_size=16)
set_row_height(ws3, 1, 40)

# ── Subject summary table ─────────────────────────────────
subj_hdrs = ["Subject", "Students", "Avg Score", "Highest", "Lowest", "Pass Rate"]
for ci, h in enumerate(subj_hdrs, 1):
    make_header_cell(ws3, ws3.cell(row=2, column=ci).coordinate, h)
    set_column_width(ws3, get_column_letter(ci), 15)
set_row_height(ws3, 2, 24)

subject_groups = df.groupby("subject")["score"]
chart_subjects = []
chart_averages = []

for ri, (subject, scores_in_subj) in enumerate(subject_groups, start=3):
    bg       = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
    avg      = scores_in_subj.mean()
    high     = scores_in_subj.max()
    low      = scores_in_subj.min()
    n        = len(scores_in_subj)
    pass_pct = f"{(scores_in_subj >= 60).sum() / n * 100:.0f}%"

    make_data_cell(ws3, f"A{ri}", subject,       bg, align="left")
    make_data_cell(ws3, f"B{ri}", n,              bg)
    make_data_cell(ws3, f"C{ri}", f"{avg:.1f}",   bg)
    make_data_cell(ws3, f"D{ri}", int(high),       bg)
    make_data_cell(ws3, f"E{ri}", int(low),        bg)
    make_data_cell(ws3, f"F{ri}", pass_pct,        bg)
    set_row_height(ws3, ri, 20)

    chart_subjects.append(subject)
    chart_averages.append(round(avg, 1))

# ── Bar Chart: Average Score by Subject ──────────────────
# Write chart source data in columns H and I
ws3["H2"].value = "Subject"
ws3["I2"].value = "Avg Score"
for i, (subj, avg) in enumerate(zip(chart_subjects, chart_averages)):
    ws3.cell(row=3 + i, column=8).value = subj
    ws3.cell(row=3 + i, column=9).value = avg

bar = BarChart()
bar.type               = "col"    # vertical bars
bar.title              = "Average Score by Subject"
bar.y_axis.title       = "Score"
bar.y_axis.scaling.min = 0
bar.y_axis.scaling.max = 100
bar.style              = 10
bar.width              = 18
bar.height             = 12

data_ref = Reference(ws3, min_col=9, min_row=2,
                     max_row=2 + len(chart_subjects))
cats_ref = Reference(ws3, min_col=8, min_row=3,
                     max_row=2 + len(chart_subjects))
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
ws3.add_chart(bar, "A10")

# ── Grade-Level Comparison Table ─────────────────────────
ws3["A28"].value = "Grade Level Comparison"
ws3["A28"].font  = Font(name="Arial", size=13, bold=True, color="2C3E50")

for ci, h in enumerate(["Grade Level", "Students", "Avg Score", "Pass Rate"], 1):
    make_header_cell(ws3, ws3.cell(row=29, column=ci).coordinate, h)

grade_groups = df.groupby("grade_level")["score"]
for ri, (grade, g_scores) in enumerate(grade_groups, start=30):
    bg       = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
    avg      = g_scores.mean()
    n        = len(g_scores)
    pass_pct = f"{(g_scores >= 60).sum() / n * 100:.0f}%"
    make_data_cell(ws3, f"A{ri}", grade,        bg, align="left")
    make_data_cell(ws3, f"B{ri}", n,             bg)
    make_data_cell(ws3, f"C{ri}", f"{avg:.1f}",  bg)
    make_data_cell(ws3, f"D{ri}", pass_pct,      bg)
    set_row_height(ws3, ri, 20)


# ============================================================
#  SHEET 4 — Attendance Report
# ============================================================
print("🔨 Building Sheet 4: Attendance Report...")

ws4 = wb.create_sheet("Attendance Report")
ws4.sheet_view.showGridLines = False

ws4.merge_cells("A1:E1")
make_header_cell(ws4, "A1", "🗓️  Attendance Report",
                 bg_color="1A252F", text_color="F1C40F", font_size=16)
set_row_height(ws4, 1, 40)

att_hdrs   = ["ID", "Name", "Attendance %", "Status", "Progress Bar"]
att_widths = [10,    22,     15,             22,        20]
for ci, (h, w) in enumerate(zip(att_hdrs, att_widths), 1):
    make_header_cell(ws4, ws4.cell(row=2, column=ci).coordinate, h)
    set_column_width(ws4, get_column_letter(ci), w)
set_row_height(ws4, 2, 24)

# Sort students by attendance — highest first
att_df = df.sort_values("attendance_pct", ascending=False)

for ri, (_, row) in enumerate(att_df.iterrows(), start=3):
    bg  = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
    att = int(row["attendance_pct"])

    # Colour-coded status label
    if att >= 90:
        status = "✅ Excellent"
    elif att >= 75:
        status = "⚠️  Satisfactory"
    else:
        status = "❌ Poor"

    make_data_cell(ws4, f"A{ri}", row["student_id"], bg)
    make_data_cell(ws4, f"B{ri}", row["name"],        bg, align="left")
    make_data_cell(ws4, f"C{ri}", att,                bg)
    make_data_cell(ws4, f"D{ri}", status,             bg)
    make_data_cell(ws4, f"E{ri}", att,                bg)   # used for progress bar
    set_row_height(ws4, ri, 18)

# Blue data bars on the Attendance % column
end_row = 2 + len(att_df)
ws4.conditional_formatting.add(
    f"C3:C{end_row}",
    DataBarRule(start_type="num", start_value=0,
                end_type="num",   end_value=100, color="3498DB")
)
# Green data bars on the Progress Bar column
ws4.conditional_formatting.add(
    f"E3:E{end_row}",
    DataBarRule(start_type="num", start_value=0,
                end_type="num",   end_value=100, color="27AE60")
)


# ── Save the dashboard ────────────────────────────────────
output_file = "student_dashboard.xlsx"
wb.save(output_file)

print("\n✅  Dashboard created successfully!")
print(f"📄  File saved → {output_file}")
print("\nSheets in the dashboard:")
print("  📋  Sheet 1 — Student Records")
print("  📊  Sheet 2 — Summary Statistics")
print("  📈  Sheet 3 — Subject Analysis")
print("  🗓️   Sheet 4 — Attendance Report")
